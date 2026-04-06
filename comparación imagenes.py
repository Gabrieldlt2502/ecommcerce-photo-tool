import os
import tempfile
from io import BytesIO
from pathlib import Path
from tkinter import Tk, filedialog

import requests
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =========================
# CONFIGURACIÓN
# =========================
SHEET_NAME = None              # None = hoja activa del archivo base
OUTPUT_SHEET_NAME = "Comparativo"

START_ROW = 2

# Base de datos
COL_SKU = "A"
COL_NUM = "N"

# Links ML en la base
LINK_COLUMNS = ["D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]

# Imágenes locales a buscar
LOCAL_SUFFIXES = ["_01", "_02"]

SUPPORTED_EXTENSIONS = [".jpg", ".jpeg", ".png", ".webp", ".bmp"]

# Tamaños
IMG_WIDTH = 120
IMG_HEIGHT = 120
ROW_HEIGHT_IMG = 95
ROW_HEIGHT_TEXT = 22
ROW_HEIGHT_LINKS = 45
ROW_HEIGHT_BLANK = 10

# Columnas en hoja comparativa
LABEL_COL = 1        # A
DATA_START_COL = 2   # B

# Estilo
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
LABEL_FILL = PatternFill("solid", fgColor="EFEFEF")
YESNO_FILL = PatternFill("solid", fgColor="FFF2CC")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def seleccionar_archivo_excel():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    archivo = filedialog.askopenfilename(
        title="Selecciona el Excel con SKUs",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    root.destroy()
    return archivo


def seleccionar_carpeta():
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    carpeta = filedialog.askdirectory(
        title="Selecciona la carpeta con las imágenes"
    )
    root.destroy()
    return carpeta


def normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def normalizar_url(url):
    if not url:
        return None

    url = str(url).strip()
    if not url:
        return None

    # Corrige links tipo http://http2...
    url = url.replace("http://http2", "https://http2")
    url = url.replace("http://", "https://")

    # Si quieres versión grande, descomenta esta línea:
    # url = url.replace("-F.jpg", "-O.jpg")

    return url


def construir_indice_imagenes(folder_path):
    image_index = {}

    for file_name in os.listdir(folder_path):
        full_path = os.path.join(folder_path, file_name)

        if not os.path.isfile(full_path):
            continue

        ext = Path(file_name).suffix.lower()
        if ext in SUPPORTED_EXTENSIONS:
            stem = Path(file_name).stem.strip().lower()
            if stem not in image_index:
                image_index[stem] = full_path

    return image_index


def buscar_imagen_local(nombre, image_index):
    if not nombre:
        return None
    key = str(nombre).strip().lower()
    return image_index.get(key)


def descargar_imagen_a_temp(url, cache_descargas):
    if url in cache_descargas:
        return cache_descargas[url]

    try:
        response = requests.get(url, timeout=20)
        response.raise_for_status()

        img_bytes = BytesIO(response.content)
        pil_img = PILImage.open(img_bytes)
        formato = pil_img.format.lower() if pil_img.format else "jpg"

        if formato == "jpeg":
            ext = ".jpg"
        elif formato in ["png", "jpg", "bmp", "gif", "webp"]:
            ext = "." + formato
        else:
            ext = ".jpg"

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        pil_img.save(temp_file.name)
        temp_file.close()

        cache_descargas[url] = temp_file.name
        return temp_file.name

    except Exception:
        cache_descargas[url] = None
        return None


def insertar_imagen(ws, image_path, row, col):
    img = XLImage(image_path)
    img.width = IMG_WIDTH
    img.height = IMG_HEIGHT
    cell_address = ws.cell(row=row, column=col).coordinate
    ws.add_image(img, cell_address)


def escribir_celda(ws, row, col, value="", fill=None, bold=False, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER

    if fill:
        cell.fill = fill

    if bold:
        cell.font = Font(bold=True)

    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    return cell


def preparar_hoja_comparativo(wb):
    if OUTPUT_SHEET_NAME in wb.sheetnames:
        del wb[OUTPUT_SHEET_NAME]

    ws = wb.create_sheet(OUTPUT_SHEET_NAME)
    ws.freeze_panes = "B2"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18

    for col_idx in range(3, 30):  # C:AD aprox
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 22

    return ws


def obtener_links_de_fila(ws_base, row):
    links = []
    for col in LINK_COLUMNS:
        valor = ws_base[f"{col}{row}"].value
        texto = normalizar_texto(valor)
        if texto:
            links.append(texto)
    return links


def main():
    print("Selecciona el Excel con SKUs...")
    excel_path = seleccionar_archivo_excel()
    if not excel_path:
        print("No seleccionaste el Excel con SKUs.")
        return

    print("Selecciona la carpeta con las imágenes...")
    images_folder = seleccionar_carpeta()
    if not images_folder:
        print("No seleccionaste la carpeta con las imágenes.")
        return

    print("Construyendo índice de imágenes locales...")
    image_index = construir_indice_imagenes(images_folder)
    print(f"Imágenes locales indexadas: {len(image_index)}")

    print("Abriendo Excel...")
    wb = load_workbook(excel_path)
    ws_base = wb[SHEET_NAME] if SHEET_NAME and SHEET_NAME in wb.sheetnames else wb.active
    ws_out = preparar_hoja_comparativo(wb)

    temp_files = []
    cache_descargas = {}

    ml_insertadas = 0
    locales_insertadas = 0
    ml_no_encontradas = []
    locales_no_encontradas = []
    bloques_creados = 0

    max_row = ws_base.max_row
    out_row = 1

    print(f"Filas a revisar en base: {max_row - START_ROW + 1}")
    print("Iniciando proceso...\n")

    for row in range(START_ROW, max_row + 1):
        sku = normalizar_texto(ws_base[f"{COL_SKU}{row}"].value)
        num = normalizar_texto(ws_base[f"{COL_NUM}{row}"].value)
        links = obtener_links_de_fila(ws_base, row)

        # Saltar filas completamente vacías
        if not sku and not num and not links:
            continue

        bloques_creados += 1
        print(f"Procesando fila base {row}/{max_row} | SKU={sku} | num={num}")

        # nombres locales
        local_names = []
        if num:
            for suffix in LOCAL_SUFFIXES:
                local_names.append(f"{num}{suffix}")

        max_items = max(len(links), len(local_names), 2)

        # -------------------------
        # Fila 1 del bloque: SKU + encabezados link
        # -------------------------
        escribir_celda(ws_out, out_row, LABEL_COL, "SKU / Links", fill=HEADER_FILL, bold=True)
        escribir_celda(ws_out, out_row, DATA_START_COL, sku, fill=HEADER_FILL, bold=True)

        for i in range(max_items):
            col = DATA_START_COL + 1 + i  # C en adelante
            if i < len(links):
                escribir_celda(ws_out, out_row, col, f"link {i+1}", fill=HEADER_FILL, bold=True, center=True)
            else:
                escribir_celda(ws_out, out_row, col, "", fill=HEADER_FILL)

        ws_out.row_dimensions[out_row].height = ROW_HEIGHT_TEXT
        out_row += 1

        # -------------------------
        # Fila 2 del bloque: fotos ML
        # -------------------------
        escribir_celda(ws_out, out_row, LABEL_COL, "Fotos ML", fill=LABEL_FILL, bold=True)
        escribir_celda(ws_out, out_row, DATA_START_COL, "")

        for i in range(max_items):
            col = DATA_START_COL + 1 + i
            escribir_celda(ws_out, out_row, col, "")

        ws_out.row_dimensions[out_row].height = ROW_HEIGHT_IMG

        for i, raw_link in enumerate(links):
            col = DATA_START_COL + 1 + i
            url = normalizar_url(raw_link)

            if not url:
                ml_no_encontradas.append(f"Fila base {row} | SKU={sku} | link inválido: {raw_link}")
                continue

            print(f"   Descargando ML {i+1}/{len(links)}...")
            temp_img_path = descargar_imagen_a_temp(url, cache_descargas)

            if temp_img_path:
                try:
                    insertar_imagen(ws_out, temp_img_path, out_row, col)
                    if temp_img_path not in temp_files:
                        temp_files.append(temp_img_path)
                    ml_insertadas += 1
                except Exception as e:
                    ml_no_encontradas.append(f"Fila base {row} | SKU={sku} | error insertando {url} | {e}")
            else:
                ml_no_encontradas.append(f"Fila base {row} | SKU={sku} | no descargada: {url}")

        out_row += 1

        # -------------------------
        # Fila 3 del bloque: links ML (texto)
        # -------------------------
        escribir_celda(ws_out, out_row, LABEL_COL, "Links ML", fill=LABEL_FILL, bold=True)
        escribir_celda(ws_out, out_row, DATA_START_COL, "")

        for i in range(max_items):
            col = DATA_START_COL + 1 + i
            if i < len(links):
                escribir_celda(ws_out, out_row, col, links[i], center=False)
            else:
                escribir_celda(ws_out, out_row, col, "")

        ws_out.row_dimensions[out_row].height = ROW_HEIGHT_LINKS
        out_row += 1

        # -------------------------
        # Fila 4 del bloque: nombres locales
        # -------------------------
        escribir_celda(ws_out, out_row, LABEL_COL, "Locales", fill=LABEL_FILL, bold=True)
        escribir_celda(ws_out, out_row, DATA_START_COL, num if num else "")

        for i in range(max_items):
            col = DATA_START_COL + 1 + i
            if i < len(local_names):
                escribir_celda(ws_out, out_row, col, local_names[i], center=True)
            else:
                escribir_celda(ws_out, out_row, col, "")

        ws_out.row_dimensions[out_row].height = ROW_HEIGHT_TEXT
        out_row += 1

        # -------------------------
        # Fila 5 del bloque: fotos locales
        # -------------------------
        escribir_celda(ws_out, out_row, LABEL_COL, "Fotos locales", fill=LABEL_FILL, bold=True)
        escribir_celda(ws_out, out_row, DATA_START_COL, "")

        for i in range(max_items):
            col = DATA_START_COL + 1 + i
            escribir_celda(ws_out, out_row, col, "")

        ws_out.row_dimensions[out_row].height = ROW_HEIGHT_IMG

        for i, local_name in enumerate(local_names):
            col = DATA_START_COL + 1 + i
            local_path = buscar_imagen_local(local_name, image_index)

            if local_path:
                try:
                    insertar_imagen(ws_out, local_path, out_row, col)
                    locales_insertadas += 1
                except Exception as e:
                    locales_no_encontradas.append(
                        f"Fila base {row} | SKU={sku} | {local_name} | error insertando | {e}"
                    )
            else:
                locales_no_encontradas.append(
                    f"Fila base {row} | SKU={sku} | no encontrada local: {local_name}"
                )

        out_row += 1

        # -------------------------
        # Fila 6 del bloque: revisión
        # -------------------------
        escribir_celda(ws_out, out_row, LABEL_COL, "Revisión", fill=YESNO_FILL, bold=True)
        escribir_celda(ws_out, out_row, DATA_START_COL, "Sí / No", fill=YESNO_FILL, bold=True, center=True)

        for i in range(max_items):
            col = DATA_START_COL + 1 + i
            escribir_celda(ws_out, out_row, col, "", fill=YESNO_FILL)

        ws_out.row_dimensions[out_row].height = ROW_HEIGHT_TEXT
        out_row += 1

        # -------------------------
        # Fila 7 del bloque: espacio en blanco
        # -------------------------
        ws_out.row_dimensions[out_row].height = ROW_HEIGHT_BLANK
        out_row += 1

    original = Path(excel_path)
    output_path = original.with_name(f"{original.stem}_comparativo_imagenes.xlsx")

    print("\nGuardando archivo...")
    wb.save(output_path)

    print("Limpiando archivos temporales...")
    for temp_path in temp_files:
        try:
            os.remove(temp_path)
        except Exception:
            pass

    print("\nListo.")
    print(f"Archivo guardado en: {output_path}")
    print(f"Bloques creados: {bloques_creados}")
    print(f"Imágenes ML insertadas: {ml_insertadas}")
    print(f"Imágenes locales insertadas: {locales_insertadas}")
    print(f"ML no encontradas: {len(ml_no_encontradas)}")
    print(f"Locales no encontradas: {len(locales_no_encontradas)}")

    if ml_no_encontradas:
        print("\nPrimeras ML no encontradas:")
        for item in ml_no_encontradas[:20]:
            print(" -", item)

    if locales_no_encontradas:
        print("\nPrimeras locales no encontradas:")
        for item in locales_no_encontradas[:20]:
            print(" -", item)


if __name__ == "__main__":
    main()